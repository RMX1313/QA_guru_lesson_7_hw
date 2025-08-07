from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from io import TextIOWrapper



def test_zip():

    with ZipFile('zip_file.zip', 'w') as zipf:
        zipf.write('test.csv')
        zipf.write('test.xlsx')
        zipf.write('test.pdf')


    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('test.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            print(reader.pages[0].extract_text())
            assert "Testpdf" in reader.pages[0].extract_text()

    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('test.xlsx') as xlsx_file:
            woorkbook = load_workbook(xlsx_file)
            sheet = woorkbook.active
            print(sheet.cell(row=2, column=1).value)
            assert 'StartDate' in sheet.cell(row=1, column=2).value

    with ZipFile('zip_file.zip', 'r') as zipf:
        with zipf.open('test.csv') as csv_file:
            text_file = TextIOWrapper(csv_file, encoding='utf-8-sig')
            content = text_file.read()
            assert "trailhead9.ub20k5i9t8ou@example.com" in content